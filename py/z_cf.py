##
from datetime import datetime as dt
import pathlib
import itertools

import numpy as np
import openpyxl as pyxl
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from persiantools.jdatetime import JalaliDate

from py import z_ns

# Shortened Namespaces
po = z_ns.Portfolio()
dirs = z_ns.Dirs()
gf = z_ns.GlobalFiles()
apw = z_ns.AdjPricesWithDates()

class Portfolio:
    def __init__(self):
        self.portfo = pd.DataFrame(columns = [po.id, po.ticker, po.tradeDate,
                                              po.jMonthsReturns,
                                              po.lastMonthReturns])

    def add_to_portfo(self,
                      ids_list,
                      tikers_list,
                      trade_date,
                      j_month_returns_list):
        df = pd.DataFrame({
                po.id            : list(ids_list),
                po.ticker        : list(tikers_list),
                po.jMonthsReturns: list(j_month_returns_list)})
        df[po.tradeDate] = trade_date
        self.portfo = self.portfo.append(df)

    def remove_by_column_value(self, column_name, value):
        self.portfo = self.portfo[self.portfo[column_name].ne(value)]

    def update_last_month_returns(self, trade_df, end_of_month_date):
        df = trade_df.loc[trade_df[po.date].eq(end_of_month_date)]
        df = df[[po.id, po.lastMonthReturns]]
        self.portfo = self.portfo.drop(columns = po.lastMonthReturns)
        self.portfo = self.portfo.merge(df, how = 'left')

    def get_last_month_eq_weight_returns(self, trade_df, end_of_month_date):
        self.update_last_month_returns(trade_df = trade_df,
                                       end_of_month_date = end_of_month_date)
        return self.portfo[po.lastMonthReturns].mean()

def save_df_to_xl(df: pd.DataFrame,
                  pn_suff_less: pathlib.Path,
                  index: bool = False,
                  header: bool = True,
                  max_col_length: int = 40):
    wb = pyxl.Workbook()
    ws = wb.active
    df = df.fillna(value = '')
    for r in dataframe_to_rows(df, index = index, header = header):
        ws.append(r)
    panes = index * ws['A'] + header * ws[1]
    for cell in panes:
        cell.style = 'Pandas'
    for column in ws.columns:
        length = max(len(str(cell.value)) for cell in column)
        length = length if length <= max_col_length else max_col_length
        ws.column_dimensions[column[0].column_letter].width = length
    wb.save(pn_suff_less.resolve().with_suffix(z_ns.xlsuf))
    wb.close()

def return_clusters_indices(iterable_obj, clustersize=100):
    intdiv = len(iterable_obj) // clustersize
    clusters_indices = [x * clustersize for x in range(0, intdiv + 1)]
    if len(clusters_indices) > 1:
        if clusters_indices[-1] != len(iterable_obj):
            clusters_indices.append(
                    clusters_indices[-1] + len(iterable_obj) % clustersize)
    else:
        clusters_indices = [0, len(iterable_obj)]
        if clusters_indices == [0, 0]:
            clusters_indices = [0]
    print(clusters_indices)
    return clusters_indices

def find_nth_workday_jdate_each_month(filled_adjprices, nth_day):
    df = filled_adjprices
    df[apw.nthWD] = nth_day
    s1 = df.groupby(apw.jMonth)[apw.jDate].unique()
    s1 = s1.apply(lambda x: np.sort(x)[nth_day])
    df1 = s1.to_frame().reset_index()
    df1 = df1.rename(columns = {apw.jDate: apw.nthWDJDate})
    df = df.merge(df1, how = "left")
    df2 = df[df[apw.jDate].eq(df[apw.nthWDJDate])][
        [apw.date, apw.nthWDJDate]].drop_duplicates()
    df2 = df2.rename(columns = {apw.date: apw.nthWDDate})
    df = df.merge(df2, how = "left")
    return df

def find_wd_after_skipping_mmonth_and_pdays(df_with_nth_wday_each_mon,
                                            m_months_skip,
                                            p_days_to_skip):
    df = df_with_nth_wday_each_mon
    df[apw.mMonthSkip] = m_months_skip
    df[apw.pDaysSkip] = p_days_to_skip
    df1 = df[[apw.date, apw.nthWDDate]].drop_duplicates()
    df1['Date+Skip'] = df1[apw.nthWDDate].apply(lambda x: x + relativedelta(
        months = m_months_skip,
        days = p_days_to_skip))
    df1[apw.afterSkipWDDate] = df1['Date+Skip'].apply(lambda x:
    df1[df1[apw.date].ge(x)][apw.date].min())
    df1[apw.afterSkipWDJDate] = df1[apw.afterSkipWDDate].apply(lambda x: str(
        JalaliDate.to_jalali(x)) if not pd.isna(x) else x)
    df = df.merge(df1, how = 'left')
    return df

def find_and_save_nth_day_and_skip_dates(filled_adjprices,
                                         nth_day,
                                         m_months_skip,
                                         p_days_to_skip):
    df = find_nth_workday_jdate_each_month(filled_adjprices, nth_day)
    df = find_wd_after_skipping_mmonth_and_pdays(df,
                                                 m_months_skip,
                                                 p_days_to_skip)
    df = df[df[apw.date].isin(df[apw.nthWDDate]) | df[apw.date].isin(
            df[apw.afterSkipWDDate])]
    assert not df.duplicated(subset = [apw.id, apw.date]).all()
    df = df.reset_index(drop = True)
    feather_pn = make_relavant_data_feather_pathname(nth_day,
                                                     m_months_skip,
                                                     p_days_to_skip)
    df.to_feather(feather_pn)
    return df

def make_relavant_data_feather_pathname(formation_workday_each_month,
                                        months_to_skip,
                                        days_to_skip):
    return dirs.cache / f'{formation_workday_each_month}-{months_to_skip}-' \
                        f'{days_to_skip}'

def get_adjprices():
    return pd.read_parquet(pathlib.Path(gf.adjPricesPrq).resolve())

def compute_returns(prices_df,
                    new_column_name,
                    groupby_list,
                    which_col_returns,
                    how_many_periods):
    prices_df[new_column_name] = prices_df.groupby(groupby_list)[
        which_col_returns].pct_change(periods = how_many_periods)
    return prices_df

def compute_bins_numbers(prices_df, groupby_list, which_col, how_many_bins):
    prices_df.loc[prices_df[which_col].notna(), apw.binNo] = \
        prices_df[prices_df[which_col].notna()].groupby(groupby_list)[
            which_col].apply(lambda x: 1 + pd.cut(x,
                                                  bins = how_many_bins,
                                                  labels = False))
    return prices_df

def build_all_possible_configs(eval_day_skip_m_skip_d,
                               from_month_to_month,
                               eval_months,
                               holding_months,
                               qcuts):
    """ examples:
        evalSkipMonthSkipDay = [(0, 0, 0), (0, 1, 0), (0, 0, 7)]
        fromMonthToMonth = [(138001, 139909), (138001, 139712), (138001,
        139512),(138001, 139012), ]
        evalMonths = [3, 6, 9, 11, 12]
        holdingMonths = [3, 6, 9, 12]
        qCut = [3, 5, 6, 10]
    """
    rs = z_ns.RSSParams()
    all_vals = [eval_day_skip_m_skip_d, from_month_to_month, eval_months,
                holding_months, qcuts]
    all_sts = list(itertools.product(*all_vals))
    all_configs = []
    for el in all_sts:
        config = {
                rs.nthWDayEachMonthForEval: el[0][0],
                rs.mMonthSkip             : el[0][1],
                rs.pDaysSkip              : el[0][2],
                rs.fromJMonth             : el[1][0],
                rs.toJMonth               : el[1][1],
                rs.evalMonths             : el[2],
                rs.holdingMonths          : el[3],
                rs.qCut                   : el[4], }
        print(config)
        all_configs.append(config)
    print(len(all_configs))
    return all_configs

##
def lookup_by_unique_key_val_in_df(df, lookupcol, lookupval, resultcol):
    o = list(df[df[lookupcol].eq(lookupval)][resultcol])
    o = np.unique(o)
    if len(o) == 1:
        return o[0]
    else:
        return 0

##


def main():
    pass

    ##
    df = get_adjprices()
    df1 = find_nth_workday_jdate_each_month(df, 0)
    df2 = find_wd_after_skipping_mmonth_and_pdays(df1, 1, 0)

    ##
    df3 = find_and_save_nth_day_and_skip_dates(df, 0, 1, 0)
    ##
    df = pd.DataFrame([[1, 2], [3, 4]],
                      columns = list('AB'),
                      index = ['x', 'y'])
    df2 = pd.DataFrame([[1, 2], [3, 4]],
                       columns = list('AB'),
                       index = ['x', 'y'])
    df.append(df2)
    ##
    df = pd.DataFrame({
            po.id            : list([1, 2]),
            po.ticker        : list(['a', 'b']),
            po.jMonthsReturns: list([.1, -.2])})
    df[po.tradeDate] = 139909
    df1 = pd.DataFrame({
            po.id              : list([1, 2]),
            po.ticker          : list(['a', 'b']),
            po.jMonthsReturns  : list([.1, -.2]),
            po.lastMonthReturns: [0.2, .3]})
    df1[po.tradeDate] = 139909
    df2 = df1.append(df)
    df2
    ##
    port = Portfolio()
    port.add_to_portfo(df[po.id], df[po.ticker], 139806, df[po.jMonthsReturns])
    port.portfo
    ##
    port.update_last_month_returns(df1, 139909)
    port.portfo  ##
